import sys
import requests
from bs4 import BeautifulSoup
import re
from urllib.parse import urljoin, urlparse, parse_qs

# PyQt5 Imports
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QLineEdit, QProgressBar, QLabel, QAbstractItemView, QFileDialog, QMessageBox
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from PyQt5.QtGui import QColor, QFont, QIcon

class CrawlWorker(QThread):
    # Signals to communicate with the GUI thread
    progress_signal = pyqtSignal(int, int)  # current_page, total_pages
    finished_signal = pyqtSignal(list)      # list of crawled stocks
    error_signal = pyqtSignal(str)          # error message

    def __init__(self):
        super().__init__()
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        }

    def run(self):
        main_url = "https://finance.naver.com/sise/sise_index.naver?code=KPI200"
        base_url = "https://finance.naver.com"
        
        try:
            res = requests.get(main_url, headers=self.headers)
            res.raise_for_status()
        except Exception as e:
            self.error_signal.emit(f"메인 페이지 연결 실패: {str(e)}")
            return

        # Find iframe URL dynamically
        html = res.content.decode('euc-kr', errors='replace')
        soup = BeautifulSoup(html, 'html.parser')
        iframe = soup.find('iframe', src=re.compile(r'/sise/entryJongmok\.naver'))
        
        if iframe:
            iframe_src = iframe.get('src')
            iframe_url = urljoin(base_url, iframe_src)
        else:
            iframe_url = "https://finance.naver.com/sise/entryJongmok.naver?type=KPI200"

        # Determine total pages
        first_page_url = f"{iframe_url}&page=1"
        try:
            first_res = requests.get(first_page_url, headers=self.headers)
            first_res.raise_for_status()
        except Exception as e:
            self.error_signal.emit(f"첫 페이지 연결 실패: {str(e)}")
            return

        first_html = first_res.content.decode('euc-kr', errors='replace')
        first_soup = BeautifulSoup(first_html, 'html.parser')

        detected_total_pages = 1
        navi_table = first_soup.find('table', class_='Nnavi')
        if navi_table:
            links = navi_table.find_all('a')
            for link in links:
                href = link.get('href', '')
                parsed_href = urlparse(href)
                query_params = parse_qs(parsed_href.query)
                page_vals = query_params.get('page', [])
                if page_vals:
                    try:
                        page_num = int(page_vals[0])
                        if page_num > detected_total_pages:
                            detected_total_pages = page_num
                    except ValueError:
                        pass

        stocks = []
        
        # Scrape all pages
        for page in range(1, detected_total_pages + 1):
            # Emit progress signal (current page, total pages)
            self.progress_signal.emit(page, detected_total_pages)
            
            page_url = f"{iframe_url}&page={page}"
            try:
                page_res = requests.get(page_url, headers=self.headers)
                page_res.raise_for_status()
            except Exception as e:
                self.error_signal.emit(f"{page}페이지 수집 실패: {str(e)}")
                break

            page_html = page_res.content.decode('euc-kr', errors='replace')
            page_soup = BeautifulSoup(page_html, 'html.parser')
            
            table = page_soup.find('table', class_='type_1')
            if not table:
                break
                
            rows = table.find_all('tr')
            page_stocks_count = 0
            
            for row in rows:
                name_cell = row.find('td', class_='ctg')
                if not name_cell:
                    continue
                    
                link = name_cell.find('a')
                if not link:
                    continue
                    
                stock_name = link.get_text(strip=True)
                href = link.get('href')
                parsed_href = urlparse(href)
                query_params = parse_qs(parsed_href.query)
                stock_code = query_params.get('code', [''])[0]
                
                cells = row.find_all('td')
                if len(cells) < 7:
                    continue
                    
                price = cells[1].get_text(strip=True)
                
                change_cell = cells[2]
                direction_span = change_cell.find(class_='blind')
                direction = direction_span.get_text(strip=True) if direction_span else ""
                change_val_span = change_cell.find('span', class_=re.compile(r'tah|red|nv'))
                change_value = change_val_span.get_text(strip=True) if change_val_span else ""
                
                rate_span = cells[3].find('span')
                rate = rate_span.get_text(strip=True) if rate_span else cells[3].get_text(strip=True)
                
                volume = cells[4].get_text(strip=True)
                amount = cells[5].get_text(strip=True)
                market_cap = cells[6].get_text(strip=True)
                
                change_sign = ""
                if "상승" in direction or "상한" in direction:
                    change_sign = "▲ "
                elif "하락" in direction or "하한" in direction:
                    change_sign = "▼ "
                elif "보합" in direction:
                    change_sign = "  "
                    
                net_change = f"{change_sign}{change_value}" if change_value else "0"
                
                stocks.append({
                    "rank": len(stocks) + 1,
                    "code": stock_code,
                    "name": stock_name,
                    "price": price,
                    "net_change": net_change,
                    "rate": rate,
                    "volume": volume,
                    "amount": amount,
                    "market_cap": market_cap
                })
                page_stocks_count += 1
                
            if page_stocks_count == 0:
                break
                
        self.finished_signal.emit(stocks)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.stocks_data = []
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("KOSPI 200 편입 종목 실시간 크롤러")
        self.resize(1000, 750)
        
        # QSS Stylesheet for modern and premium styling
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f4f6f9;
            }
            QWidget {
                font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
                font-size: 13px;
                color: #333333;
            }
            QLabel#titleLabel {
                font-size: 20px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px 0px;
            }
            QPushButton#btnCrawl {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 18px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton#btnCrawl:hover {
                background-color: #2980b9;
            }
            QPushButton#btnCrawl:pressed {
                background-color: #1f618d;
            }
            QPushButton#btnCrawl:disabled {
                background-color: #bdc3c7;
            }
            QPushButton#btnExport {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 18px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton#btnExport:hover {
                background-color: #27ae60;
            }
            QPushButton#btnExport:pressed {
                background-color: #1e8449;
            }
            QPushButton#btnExport:disabled {
                background-color: #bdc3c7;
            }
            QLineEdit#searchBar {
                border: 1px solid #ccd1d9;
                border-radius: 5px;
                padding: 7px 12px;
                background-color: white;
                font-size: 13px;
            }
            QLineEdit#searchBar:focus {
                border-color: #3498db;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #e6e8eb;
                border-radius: 6px;
                gridline-color: #f1f3f5;
            }
            QHeaderView::section {
                background-color: #ecf0f1;
                color: #34495e;
                padding: 8px;
                border: none;
                border-bottom: 2px solid #bdc3c7;
                font-weight: bold;
            }
            QTableWidget::item:selected {
                background-color: #e8f4fc;
                color: #333333;
            }
            QProgressBar {
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                background-color: #e2e6ea;
                height: 18px;
            }
            QProgressBar::chunk {
                background-color: #2ecc71;
                border-radius: 4px;
            }
            QLabel#statusLabel {
                color: #7f8c8d;
                font-size: 12px;
            }
        """)

        # Central Widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # Header Title
        title_label = QLabel("KOSPI 200 편입 종목 크롤러")
        title_label.setObjectName("titleLabel")
        main_layout.addWidget(title_label)

        # Control Panel layout (Horizontal)
        control_layout = QHBoxLayout()
        
        self.btn_crawl = QPushButton("데이터 수집 시작")
        self.btn_crawl.setObjectName("btnCrawl")
        self.btn_crawl.clicked.connect(self.start_crawling)
        control_layout.addWidget(self.btn_crawl)

        self.btn_export = QPushButton("엑셀 파일로 저장")
        self.btn_export.setObjectName("btnExport")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self.export_to_excel)
        control_layout.addWidget(self.btn_export)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        control_layout.addWidget(self.progress_bar)
        
        control_layout.addStretch()

        # Search Bar
        self.search_bar = QLineEdit()
        self.search_bar.setObjectName("searchBar")
        self.search_bar.setPlaceholderText("종목명 또는 코드 검색...")
        self.search_bar.setFixedWidth(250)
        self.search_bar.textChanged.connect(self.filter_table)
        control_layout.addWidget(self.search_bar)

        main_layout.addLayout(control_layout)

        # Table Widget
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "순위", "종목코드", "종목명", "현재가", 
            "전일비", "등락률", "거래량", "거래대금(백만)", "시가총액(억)"
        ])
        
        # Table configurations
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Set specific resize modes for columns
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents) # 순위
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents) # 종목코드
        
        main_layout.addWidget(self.table)

        # Status Bar / Label
        self.lbl_status = QLabel("준비 완료. '데이터 수집 시작' 버튼을 누르세요.")
        self.lbl_status.setObjectName("statusLabel")
        main_layout.addWidget(self.lbl_status)

    def start_crawling(self):
        self.btn_crawl.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.search_bar.clear()
        self.table.setRowCount(0)
        
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.lbl_status.setText("데이터 수집 중...")

        # Initialize and start Worker Thread
        self.worker = CrawlWorker()
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_crawl_finished)
        self.worker.error_signal.connect(self.on_crawl_error)
        self.worker.start()

    def update_progress(self, current, total):
        percentage = int((current / total) * 100)
        self.progress_bar.setValue(percentage)
        self.lbl_status.setText(f"페이지 로드 중... ({current} / {total})")

    def on_crawl_finished(self, stocks):
        self.stocks_data = stocks
        self.display_data(stocks)
        
        self.btn_crawl.setEnabled(True)
        self.btn_export.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.lbl_status.setText(f"수집 완료: 총 {len(stocks)}개 종목이 조회되었습니다.")

    def on_crawl_error(self, err_msg):
        self.btn_crawl.setEnabled(True)
        self.btn_export.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.lbl_status.setText(f"오류 발생: {err_msg}")

    def display_data(self, stocks_list):
        self.table.setRowCount(len(stocks_list))
        
        # Set fonts and colors
        font_bold = QFont()
        font_bold.setBold(True)
        
        red_color = QColor(220, 53, 69)      # Rising color (Soft Red)
        blue_color = QColor(0, 102, 204)     # Falling color (Sleek Blue)
        black_color = QColor(51, 51, 51)     # Normal/Flat color
        
        for row_idx, stock in enumerate(stocks_list):
            # 0. Rank
            rank_item = QTableWidgetItem(str(stock["rank"]))
            rank_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 0, rank_item)
            
            # 1. Code
            code_item = QTableWidgetItem(stock["code"])
            code_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_idx, 1, code_item)
            
            # 2. Name (Bold)
            name_item = QTableWidgetItem(stock["name"])
            name_item.setFont(font_bold)
            name_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setItem(row_idx, 2, name_item)
            
            # 3. Price
            price_item = QTableWidgetItem(stock["price"])
            price_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row_idx, 3, price_item)
            
            # Decide color based on change direction
            if "▲" in stock["net_change"] or "+" in stock["rate"]:
                text_color = red_color
            elif "▼" in stock["net_change"] or "-" in stock["rate"]:
                text_color = blue_color
            else:
                text_color = black_color

            # 4. Net Change
            change_item = QTableWidgetItem(stock["net_change"])
            change_item.setForeground(text_color)
            change_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row_idx, 4, change_item)
            
            # 5. Fluctuation Rate
            rate_item = QTableWidgetItem(stock["rate"])
            rate_item.setForeground(text_color)
            rate_item.setFont(font_bold)
            rate_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row_idx, 5, rate_item)
            
            # 6. Volume
            volume_item = QTableWidgetItem(stock["volume"])
            volume_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row_idx, 6, volume_item)
            
            # 7. Amount
            amount_item = QTableWidgetItem(stock["amount"])
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row_idx, 7, amount_item)
            
            # 8. Market Cap
            cap_item = QTableWidgetItem(stock["market_cap"])
            cap_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(row_idx, 8, cap_item)

    def export_to_excel(self):
        if not self.stocks_data:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 수집을 실행해 주세요.")
            return

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "엑셀 파일 저장", "KOSPI200_편입종목.xlsx", 
            "Excel Files (*.xlsx);;All Files (*)", options=options
        )

        if not file_path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "코스피200 편입종목"

            # Enable grid lines visible
            ws.views.sheetView[0].showGridLines = True

            headers = [
                "순위", "종목코드", "종목명", "현재가", 
                "전일비", "등락률", "거래량", "거래대금(백만)", "시가총액(억)"
            ]
            ws.append(headers)

            # Style Header
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='BDC3C7'),
                right=Side(style='thin', color='BDC3C7'),
                top=Side(style='thin', color='BDC3C7'),
                bottom=Side(style='thin', color='BDC3C7')
            )

            for col_idx in range(1, 10):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Fonts for data cells
            font_normal = Font(name="맑은 고딕", size=10)
            font_bold = Font(name="맑은 고딕", size=10, bold=True)
            
            red_font = Font(name="맑은 고딕", size=10, color="C0392B", bold=True)
            blue_font = Font(name="맑은 고딕", size=10, color="2980B9", bold=True)

            for stock in self.stocks_data:
                # Parse numeric values for Excel
                def clean_int(val):
                    val_str = val.replace(",", "").strip()
                    try:
                        return int(val_str)
                    except ValueError:
                        return val_str

                def clean_float_percent(val):
                    val_str = val.replace("%", "").replace("+", "").replace(",", "").strip()
                    try:
                        return float(val_str) / 100.0
                    except ValueError:
                        return val

                # Determine net change sign and value
                net_change_str = stock["net_change"]
                net_val = 0
                net_val_str = re.sub(r'[▲▼\s,]', '', net_change_str)
                try:
                    if net_val_str:
                        net_val = int(net_val_str)
                        if "▼" in net_change_str:
                            net_val = -net_val
                except ValueError:
                    net_val = net_change_str

                row_data = [
                    int(stock["rank"]),
                    stock["code"],
                    stock["name"],
                    clean_int(stock["price"]),
                    net_val,
                    clean_float_percent(stock["rate"]),
                    clean_int(stock["volume"]),
                    clean_int(stock["amount"]),
                    clean_int(stock["market_cap"])
                ]
                
                ws.append(row_data)
                curr_row = ws.max_row

                # Format and style each cell in the row
                # 1. Rank
                cell = ws.cell(row=curr_row, column=1)
                cell.font = font_normal
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

                # 2. Code
                cell = ws.cell(row=curr_row, column=2)
                cell.font = font_normal
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = '@' # Text format for stock code (keeps leading zeros)
                cell.border = thin_border

                # 3. Name
                cell = ws.cell(row=curr_row, column=3)
                cell.font = font_bold
                cell.alignment = Alignment(horizontal="left")
                cell.border = thin_border

                # 4. Price
                cell = ws.cell(row=curr_row, column=4)
                cell.font = font_normal
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
                cell.border = thin_border

                # Color decision for changes
                is_up = "▲" in stock["net_change"] or "+" in stock["rate"]
                is_down = "▼" in stock["net_change"] or "-" in stock["rate"]
                cell_font = red_font if is_up else (blue_font if is_down else font_normal)

                # 5. Net Change
                cell = ws.cell(row=curr_row, column=5)
                cell.font = cell_font
                cell.alignment = Alignment(horizontal="right")
                if isinstance(net_val, int):
                    cell.number_format = '+#,##0;-#,##0;0'
                else:
                    cell.number_format = '@'
                cell.border = thin_border

                # 6. Fluctuation Rate
                cell = ws.cell(row=curr_row, column=6)
                cell.font = cell_font
                cell.alignment = Alignment(horizontal="right")
                if isinstance(cell.value, float):
                    cell.number_format = '+0.00%;-0.00%;0.00%'
                else:
                    cell.number_format = '@'
                cell.border = thin_border

                # 7, 8, 9. Volume, Amount, Market Cap
                for col_idx in [7, 8, 9]:
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.font = font_normal
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0'
                    cell.border = thin_border

            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    cell_len = sum(2 if ord(char) > 256 else 1 for char in val_str)
                    if cell_len > max_len:
                        max_len = cell_len
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

            wb.save(file_path)
            self.lbl_status.setText(f"저장 성공: {file_path}")
            QMessageBox.information(self, "성공", f"엑셀 파일이 성공적으로 저장되었습니다.\n경로: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 파일 저장 중 오류 발생:\n{str(e)}")

    def filter_table(self):
        query = self.search_bar.text().strip().lower()
        
        # Hide rows that don't match query
        for row_idx in range(self.table.rowCount()):
            code_item = self.table.item(row_idx, 1)
            name_item = self.table.item(row_idx, 2)
            
            if code_item and name_item:
                code_text = code_item.text().lower()
                name_text = name_item.text().lower()
                
                # Check if query matches code or name
                if query in code_text or query in name_text:
                    self.table.setRowHidden(row_idx, False)
                else:
                    self.table.setRowHidden(row_idx, True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
